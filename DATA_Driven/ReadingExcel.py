import openpyxl


# file-->workbook-->sheet-->rows--->cell
file="D:\sahil_learn\Automation_testing_JAVA\Sahilselenium\ParaBanking.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet3"]

rows=sheet.max_row #count number of rows in excel sheet
cols=sheet.max_column #count number of column in excel sheet

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='   ')
    print()