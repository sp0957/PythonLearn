import openpyxl

# #for same data
# file="D:\sahil_learn\Automation_Testing_Python\DemoEMPTY.xlsx"
#
# workbook=openpyxl.load_workbook(file)
# sheet=workbook["Sheet1"]
# for r in range(1,5):
#     for c in  range(1,3):
#         sheet.cell(r,c).value="welcomw"
# workbook.save(file)

#for multiple data
file="D:\sahil_learn\Automation_Testing_Python\ALL_DATA.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet2"]
sheet.cell(1,1).value=123
sheet.cell(1,2).value="jenish"
sheet.cell(1,3).value="engineer"

sheet.cell(2,1).value=345
sheet.cell(2,2).value="sahil"
sheet.cell(1,3).value="Tester"

sheet.cell(3,1).value=345
sheet.cell(3,2).value="sp"
sheet.cell(1,3).value="engineer"


workbook.save(file)

