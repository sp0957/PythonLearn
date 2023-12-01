import  openpyxl
from openpyxl.styles import PatternFill

def getRows(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_row)

def getColumn(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rowno,coulmnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(rowno,coulmnno).value

def writeData(file,sheetName,rowno,coulmnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowno,coulmnno).value = data
    workbook.save(file)


def fillGreenColor(file,sheetName,rowno,coulmnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenfill=PatternFill(start_color='60b212',
                        end_color='60b212',
                        fill_type='solid')
    sheet.cell(rowno,coulmnno).fill=greenfill
    workbook.save(file)

def fillRedColor(file,sheetName,rowno,coulmnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redfill=PatternFill(start_color='ff0000',
                        end_color='ff0000',
                        fill_type='solid')
    sheet.cell(rowno,coulmnno).fill=redfill
    workbook.save(file)








